#!/usr/bin/env python3
"""Generate FileGram Bench Excel report with tasks, dimensions, and baseline results."""

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# Color palette & styles
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUB_HEADER_FONT = Font(name="Arial", bold=True, size=10)
CHECK_FONT = Font(name="Arial", bold=True, color="2F5496", size=11)
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)

# Score color fills
SCORE_5 = PatternFill("solid", fgColor="92D050")  # green
SCORE_4 = PatternFill("solid", fgColor="C6EFCE")  # light green
SCORE_3 = PatternFill("solid", fgColor="FFEB9C")  # yellow
SCORE_2 = PatternFill("solid", fgColor="FFC7CE")  # light red
SCORE_1 = PatternFill("solid", fgColor="FF6B6B")  # red
BEST_FILL = PatternFill("solid", fgColor="92D050")


def score_fill(val):
    if val >= 4.5:
        return SCORE_5
    if val >= 3.5:
        return SCORE_4
    if val >= 2.5:
        return SCORE_3
    if val >= 1.5:
        return SCORE_2
    return SCORE_1


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, font=None, align=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font or NORMAL_FONT
    cell.alignment = align or WRAP_ALIGN
    cell.border = THIN_BORDER
    return cell


# ============================================================
# Data
# ============================================================
tasks = [
    {
        "id": "T-01",
        "name_zh": "投资分析师工作概况摘要",
        "name_en": "Investment analyst work overview summary",
        "type": "Understand",
        "dims": ["A", "B"],
        "attrs": ["reading_strategy", "output_detail", "output_structure", "tone"],
        "workspace": "t01_workspace (Victoria, 25 files)",
        "depends": "-",
        "ops": 12,
    },
    {
        "id": "T-02",
        "name_zh": "法律案件材料梳理与时间线",
        "name_en": "Legal case materials review & timeline",
        "type": "Understand",
        "dims": ["A", "B", "E"],
        "attrs": ["reading_strategy", "output_detail", "output_structure", "tone"],
        "workspace": "t02_workspace (Adam, 24 files)",
        "depends": "-",
        "ops": 12,
    },
    {
        "id": "T-03",
        "name_zh": "个人知识库搭建",
        "name_en": "Personal knowledge base creation",
        "type": "Create",
        "dims": ["B", "C"],
        "attrs": ["output_detail", "output_structure", "directory_style", "naming", "tone"],
        "workspace": "t03_workspace (empty)",
        "depends": "-",
        "ops": 10,
    },
    {
        "id": "T-04",
        "name_zh": "会议纪要与跟进文档制作",
        "name_en": "Meeting minutes & follow-up docs",
        "type": "Create",
        "dims": ["B", "C", "F"],
        "attrs": ["output_detail", "output_structure", "directory_style", "naming", "tone"],
        "workspace": "t04_workspace (empty)",
        "depends": "-",
        "ops": 8,
    },
    {
        "id": "T-05",
        "name_zh": "混乱文件夹整理归档",
        "name_en": "Messy folder cleanup & reorganization",
        "type": "Organize",
        "dims": ["A", "C"],
        "attrs": ["reading_strategy", "directory_style", "naming", "version_strategy"],
        "workspace": "t05_workspace (mixed, 30 files)",
        "depends": "-",
        "ops": 20,
    },
    {
        "id": "T-06",
        "name_zh": "多源信息综合研究报告",
        "name_en": "Multi-source synthesis research report",
        "type": "Synthesize",
        "dims": ["A", "B", "F"],
        "attrs": ["reading_strategy", "output_detail", "output_structure", "tone", "cross_modal_behavior"],
        "workspace": "t06_workspace (Victoria AAPL, 21 files)",
        "depends": "-",
        "ops": 12,
    },
    {
        "id": "T-07",
        "name_zh": "日记与笔记综合人物画像",
        "name_en": "Diary & notes synthesis into personal profile",
        "type": "Synthesize",
        "dims": ["A", "B"],
        "attrs": ["reading_strategy", "output_detail", "output_structure", "tone"],
        "workspace": "t07_workspace (Adam personal, 22 files)",
        "depends": "-",
        "ops": 10,
    },
    {
        "id": "T-08",
        "name_zh": "季度工作总结撰写",
        "name_en": "Quarterly work summary report",
        "type": "Create",
        "dims": ["B", "C", "F"],
        "attrs": ["output_detail", "output_structure", "directory_style", "naming", "cross_modal_behavior"],
        "workspace": "t08_workspace (Victoria, 18 files)",
        "depends": "-",
        "ops": 8,
    },
    {
        "id": "T-09",
        "name_zh": "报告修订与压缩",
        "name_en": "Report revision & condensation",
        "type": "Iterate",
        "dims": ["D", "B"],
        "attrs": ["edit_strategy", "version_strategy", "output_detail", "output_structure"],
        "workspace": "t09_workspace (synthetic, 1 file)",
        "depends": "-",
        "ops": 6,
    },
    {
        "id": "T-10",
        "name_zh": "知识库内容更新与维护",
        "name_en": "Knowledge base content update & maintenance",
        "type": "Maintain",
        "dims": ["D", "C", "E"],
        "attrs": ["edit_strategy", "directory_style", "version_strategy", "naming"],
        "workspace": "t10_workspace (from T-03 output)",
        "depends": "T-03",
        "ops": 8,
    },
]

ALL_DIMS = ["A", "B", "C", "D", "E", "F"]
DIM_NAMES = {
    "A": "A: Consumption\nPattern",
    "B": "B: Production\nStyle",
    "C": "C: Organization\nPreference",
    "D": "D: Iteration\nStrategy",
    "E": "E: Work\nRhythm",
    "F": "F: Cross-Modal\nBehavior",
}

ATTRIBUTES = [
    "name",
    "role",
    "language",
    "tone",
    "output_detail",
    "working_style",
    "thoroughness",
    "documentation",
    "error_handling",
    "reading_strategy",
    "output_structure",
    "directory_style",
    "naming",
    "edit_strategy",
    "version_strategy",
    "cross_modal_behavior",
]

METHODS = [
    "full_context",
    "naive_rag",
    "eager_summarization",
    "mem0",
    "zep",
    "memos",
    "memu",
    "evermemos",
    "filegramos_simple",
]

METHOD_DISPLAY = {
    "full_context": "Full Context",
    "naive_rag": "Naive RAG",
    "eager_summarization": "Eager Summarization",
    "mem0": "Mem0",
    "zep": "Zep",
    "memos": "MemOS",
    "memu": "MemU",
    "evermemos": "EverMemOS",
    "filegramos_simple": "FileGramOS Simple",
}

METHOD_DESC = {
    "full_context": "Raw behavioral trajectories (events.json) dumped into LLM context window. No memory processing.",
    "naive_rag": "Chunk trajectory into segments, embed with text embeddings, retrieve top-k chunks per attribute query.",
    "eager_summarization": "LLM summarizes each session trajectory into NL behavioral summary, then infer profile from summaries.",
    "mem0": "Convert file operations to interaction format ('I'm reading X'), extract memories via Mem0 pipeline.",
    "zep": "Build knowledge graph with entities (files/dirs) and relationships (read/wrote/edited) with temporal ordering.",
    "memos": "Organize memories into hierarchical cells (episodic/semantic/procedural) with metadata and relationships.",
    "memu": "Extract discrete memory units with type tags (access/creation/edit) and importance scores. Includes content.",
    "evermemos": "Segment trajectories by event boundaries (context switches), organize as engram-like multi-level traces.",
    "filegramos_simple": "Three-channel deterministic feature extraction (procedural stats + semantic patterns + episodic signatures). No LLM in memory building.",
}

# Load results
import pathlib

BENCH_DIR = pathlib.Path("/Users/choiszt/Desktop/code/Synvo/FileGram/bench/test_results")

ALL_PROFILES = [
    "p1_methodical",
    "p2_thorough_reviser",
    "p3_efficient_executor",
    "p4_structured_analyst",
    "p5_balanced_organizer",
    "p6_quick_curator",
]
profiles_data = {}
for pdir in ALL_PROFILES:
    rpath = BENCH_DIR / pdir / "results.json"
    if rpath.exists():
        with open(rpath) as f:
            profiles_data[pdir] = json.load(f)

loaded_profiles = list(profiles_data.keys())
num_profiles = len(loaded_profiles)

PROFILE_DISPLAY = {
    "p1_methodical": "P1 Methodical",
    "p2_thorough_reviser": "P2 Thorough Reviser",
    "p3_efficient_executor": "P3 Efficient Executor",
    "p4_structured_analyst": "P4 Structured Analyst",
    "p5_balanced_organizer": "P5 Balanced Organizer",
    "p6_quick_curator": "P6 Quick Curator",
}

# ============================================================
# Sheet 1: Task Overview & Dimension Coverage
# ============================================================
ws1 = wb.active
ws1.title = "Tasks & Dimensions"

headers1 = [
    "Task ID",
    "Task Name (中文)",
    "Task Name (English)",
    "Type",
    *[DIM_NAMES[d] for d in ALL_DIMS],
    "Observable Attributes",
    "Workspace Bundle",
    "Depends On",
    "Expected Ops",
]
for col, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=col, value=h)
style_header_row(ws1, 1, len(headers1))

for i, t in enumerate(tasks, 2):
    style_data_cell(ws1, i, 1, BOLD_FONT, CENTER_ALIGN).value = t["id"]
    style_data_cell(ws1, i, 2).value = t["name_zh"]
    style_data_cell(ws1, i, 3).value = t["name_en"]
    style_data_cell(ws1, i, 4, BOLD_FONT, CENTER_ALIGN).value = t["type"]
    for j, d in enumerate(ALL_DIMS):
        cell = style_data_cell(ws1, i, 5 + j, CHECK_FONT, CENTER_ALIGN)
        cell.value = "✓" if d in t["dims"] else ""
    style_data_cell(ws1, i, 11).value = ", ".join(t["attrs"])
    style_data_cell(ws1, i, 12).value = t["workspace"]
    style_data_cell(ws1, i, 13, align=CENTER_ALIGN).value = t["depends"]
    style_data_cell(ws1, i, 14, align=CENTER_ALIGN).value = t["ops"]

# Coverage summary row
r = len(tasks) + 3
ws1.cell(row=r, column=1, value="Dimension Coverage Count").font = BOLD_FONT
for j, d in enumerate(ALL_DIMS):
    cnt = sum(1 for t in tasks if d in t["dims"])
    cell = ws1.cell(row=r, column=5 + j, value=cnt)
    cell.font = BOLD_FONT
    cell.alignment = CENTER_ALIGN

# Column widths
col_widths = [8, 22, 38, 12, 14, 14, 14, 14, 14, 14, 45, 35, 10, 12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height = 35

# ============================================================
# Sheet 2: Dimension Reference
# ============================================================
ws2 = wb.create_sheet("Dimension Reference")

dim_data = [
    (
        "A",
        "Consumption Pattern\n信息消费模式",
        "面对新文件如何探索、定位和理解信息",
        "L: sequential\n(逐个深读)",
        "M: targeted\n(直接搜索)",
        "R: breadth-first\n(广度浏览)",
        "UC1 主动辅助\nUC4 上下文恢复",
    ),
    (
        "B",
        "Production Style\n生产风格",
        "产出内容的格式偏好、详尽度、结构习惯",
        "L: comprehensive\n(详尽)",
        "M: balanced\n(均衡)",
        "R: minimal\n(极简)",
        "UC2 个性化默认值\nUC7 代理执行",
    ),
    (
        "C",
        "Organization Preference\n组织偏好",
        "如何管理文件系统：目录结构、命名习惯、版本策略",
        "L: deeply_nested\n(深层嵌套)",
        "M: adaptive\n(按需调整)",
        "R: flat\n(完全扁平)",
        "UC3 智能组织\nUC5 行为连续性\nUC6 冲突检测",
    ),
    (
        "D",
        "Iteration Strategy\n迭代策略",
        "修改和完善已有工作时的行为模式",
        "L: incremental\n(增量打磨)",
        "M: balanced\n(适度迭代)",
        "R: rewrite\n(大幅重写)",
        "UC5 行为连续性\nUC6 冲突检测",
    ),
    (
        "E",
        "Curation\n信息管护",
        "工作空间管理：是否主动清理不需要的文件",
        "L: selective\n(主动精简)",
        "M: pragmatic\n(适度清理)",
        "R: preservative\n(全部保留)",
        "UC3 智能组织\nUC5 行为连续性",
    ),
    (
        "F",
        "Cross-Modal Behavior\n跨模态行为",
        "是否主动使用视觉材料，如何维护图文关系",
        "L: visual-heavy\n(重视视觉)",
        "M: balanced\n(均衡)",
        "R: text-only\n(纯文本)",
        "UC2 个性化默认值\nUC7 代理执行",
    ),
]

headers2 = ["Dim", "Name / 名称", "Definition / 定义", "Left (L)", "Middle (M)", "Right (R)", "Supports Use Cases"]
for col, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=col, value=h)
style_header_row(ws2, 1, len(headers2))

for i, dd in enumerate(dim_data, 2):
    for j, val in enumerate(dd):
        cell = style_data_cell(ws2, i, j + 1)
        cell.value = val
        if j == 0:
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN

# Profile matrix
r = len(dim_data) + 4
ws2.cell(row=r, column=1, value="Profile L/M/R Matrix").font = Font(name="Arial", bold=True, size=12)
r += 1
profile_headers = ["Profile", "A", "B", "C", "D", "E", "F", "Behavioral Signature"]
for col, h in enumerate(profile_headers, 1):
    ws2.cell(row=r, column=col, value=h)
style_header_row(ws2, r, len(profile_headers))

profiles_matrix = [
    (
        "p1_methodical",
        "L",
        "L",
        "L",
        "L",
        "L",
        "M",
        "Sequential deep reader, comprehensive hierarchical output, deeply nested dirs, incremental edits, phased workflow",
    ),
    (
        "p2_thorough_reviser",
        "L",
        "L",
        "R",
        "R",
        "L",
        "M",
        "Same reading/output as p1 but flat organization, bulk rewrites — fine-grained pair with p1 (C+D differ)",
    ),
    (
        "p3_efficient_executor",
        "M",
        "R",
        "R",
        "R",
        "R",
        "R",
        "Targeted search, minimal output, flat dirs, bulk rewrites, bursty rhythm, text-only",
    ),
    (
        "p4_structured_analyst",
        "M",
        "L",
        "M",
        "L",
        "M",
        "L",
        "Targeted search but comprehensive output, adaptive organization, incremental edits, visual-heavy",
    ),
    (
        "p5_balanced_organizer",
        "R",
        "M",
        "M",
        "M",
        "M",
        "M",
        "Breadth-first browsing, balanced output, adaptive organization — the 'middle ground' profile",
    ),
    (
        "p6_quick_curator",
        "R",
        "M",
        "L",
        "R",
        "R",
        "R",
        "Breadth-first browsing, balanced output, deeply nested dirs, bulk rewrites, bursty, text-only",
    ),
]

L_FILL = PatternFill("solid", fgColor="D5E8D4")  # green-ish
M_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow-ish
R_FILL = PatternFill("solid", fgColor="F8CECC")  # red-ish

for i, pm in enumerate(profiles_matrix):
    row = r + 1 + i
    style_data_cell(ws2, row, 1, BOLD_FONT).value = pm[0]
    for j in range(6):
        cell = style_data_cell(ws2, row, 2 + j, BOLD_FONT, CENTER_ALIGN)
        cell.value = pm[1 + j]
        if pm[1 + j] == "L":
            cell.fill = L_FILL
        elif pm[1 + j] == "M":
            cell.fill = M_FILL
        elif pm[1 + j] == "R":
            cell.fill = R_FILL
    style_data_cell(ws2, row, 8).value = pm[7]

col_widths2 = [6, 25, 35, 22, 22, 22, 22]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.column_dimensions["H"].width = 70
for row_idx in range(2, 2 + len(dim_data)):
    ws2.row_dimensions[row_idx].height = 50

# ============================================================
# Sheet 3: Baseline Results (Per-attribute scores)
# ============================================================
ws3 = wb.create_sheet("Baseline Results")

# Overall comparison first
ws3.cell(row=1, column=1, value="Overall Score Comparison (avg across 16 attributes)").font = Font(
    name="Arial", bold=True, size=12
)

row = 3
overall_headers = ["Method"] + loaded_profiles + ["Overall Average"]
for col, h in enumerate(overall_headers, 1):
    ws3.cell(row=row, column=col, value=h)
style_header_row(ws3, row, len(overall_headers))

# Compute per-profile overall scores from results.json
profile_scores = {}
for profile, data in profiles_data.items():
    profile_scores[profile] = {}
    for m in METHODS:
        if m in data:
            profile_scores[profile][m] = data[m].get("avg_score", 0)
        else:
            profile_scores[profile][m] = 0

# Sort by overall average descending
method_avgs = {}
for m in METHODS:
    vals = [profile_scores[p].get(m, 0) for p in loaded_profiles]
    method_avgs[m] = round(sum(vals) / len(vals), 3) if vals else 0
sorted_methods = sorted(METHODS, key=lambda m: method_avgs[m], reverse=True)

for i, m in enumerate(sorted_methods):
    r = row + 1 + i
    style_data_cell(ws3, r, 1, BOLD_FONT).value = METHOD_DISPLAY[m]
    per_profile_vals = [profile_scores[p].get(m, 0) for p in loaded_profiles]
    for j, s in enumerate(per_profile_vals + [method_avgs[m]]):
        cell = style_data_cell(ws3, r, 2 + j, NORMAL_FONT, CENTER_ALIGN)
        cell.value = round(s, 2)
        cell.fill = score_fill(s)
        cell.number_format = "0.00"

# Best method highlight
best_row = row + 1  # filegramos_simple should be first
for col in range(1, num_profiles + 3):
    ws3.cell(row=best_row, column=col).font = Font(name="Arial", bold=True, size=10, color="006100")

# Per-attribute detail section for each profile
profiles_info = [(p, p) for p in loaded_profiles]

current_row = row + len(METHODS) + 4

for profile_key, profile_name in profiles_info:
    if profile_key not in profiles_data:
        continue

    data = profiles_data[profile_key]
    ws3.cell(row=current_row, column=1, value=f"Per-Attribute Scores: {profile_name}").font = Font(
        name="Arial", bold=True, size=12
    )
    current_row += 1

    # Ground truth row
    gt = None
    for m in METHODS:
        if m in data and "ground_truth" in data[m]:
            gt = data[m]["ground_truth"]
            break

    # Headers: Attribute | Ground Truth | method1 | method2 | ...
    attr_headers = ["Attribute", "Ground Truth"] + [METHOD_DISPLAY[m] for m in METHODS]
    for col, h in enumerate(attr_headers, 1):
        ws3.cell(row=current_row, column=col, value=h)
    style_header_row(ws3, current_row, len(attr_headers))
    current_row += 1

    for attr in ATTRIBUTES:
        style_data_cell(ws3, current_row, 1, BOLD_FONT).value = attr
        style_data_cell(ws3, current_row, 2).value = gt.get(attr, "-") if gt else "-"
        for j, m in enumerate(METHODS):
            if m in data and "judge_scores" in data[m]:
                scores = data[m]["judge_scores"]["scores"]
                if attr in scores:
                    s = scores[attr]["score"]
                    cell = style_data_cell(ws3, current_row, 3 + j, NORMAL_FONT, CENTER_ALIGN)
                    cell.value = s
                    cell.fill = score_fill(s)
        current_row += 1

    # Average row
    style_data_cell(ws3, current_row, 1, BOLD_FONT).value = "AVERAGE"
    style_data_cell(ws3, current_row, 2, BOLD_FONT).value = "-"
    for j, m in enumerate(METHODS):
        if m in data:
            avg = data[m].get("avg_score", 0)
            cell = style_data_cell(ws3, current_row, 3 + j, BOLD_FONT, CENTER_ALIGN)
            cell.value = round(avg, 2)
            cell.fill = score_fill(avg)
            cell.number_format = "0.00"
    current_row += 3

# Column widths for sheet 3
ws3.column_dimensions["A"].width = 22
for i in range(2, 2 + num_profiles):
    ws3.column_dimensions[get_column_letter(i)].width = 20
ws3.column_dimensions[get_column_letter(2 + num_profiles)].width = 16  # Overall Average col
# Per-attribute detail sections use Attribute | Ground Truth | methods...
for i in range(1, 3 + len(METHODS)):
    if i == 1:
        ws3.column_dimensions[get_column_letter(i)].width = 22
    elif i == 2:
        ws3.column_dimensions[get_column_letter(i)].width = 22
    else:
        ws3.column_dimensions[get_column_letter(i)].width = 16

# ============================================================
# Sheet 4: Baseline Methods & Prompts
# ============================================================
ws4 = wb.create_sheet("Baseline Prompts")

ws4.cell(row=1, column=1, value="Memory Baseline Methods & Prompt Templates").font = Font(
    name="Arial", bold=True, size=12
)

headers4 = ["Method", "Description", "Memory Representation", "Prompt Structure (p1_methodical example)"]
row = 3
for col, h in enumerate(headers4, 1):
    ws4.cell(row=row, column=col, value=h)
style_header_row(ws4, row, len(headers4))

prompt_structures = {
    "full_context": "System: 'You are analyzing file-system behavioral trajectories...'\n\nProvides ALL raw events sequentially:\n[1] FS Snapshot: 26 files...\n[2] Read file: AAPL_Coverage_Initiation_Notes.md...\n[3] Context switch: ... -> ...\n...(all 132 events listed)\n\nAsks: 'Infer profile for: name, role, language, ...'",
    "naive_rag": "System: 'You are analyzing file-system behavioral traces...'\n\nRetrieves top-k chunks per attribute:\n### For attribute 'name':\n  [Task T-01, chunk 0]: [1] FS Snapshot...\n  [Task T-01, chunk 1]: [1] Context switch...\n### For attribute 'reading_strategy':\n  [Task T-01, chunk 0]: ...\n\nAsks: 'Infer profile for each attribute'",
    "eager_summarization": "System: 'You are analyzing behavioral summaries...'\n\nProvides LLM-generated NL summary per session:\n=== Session 1 ===\n'The user demonstrates a highly systematic and iterative approach to file exploration... Their exploration begins with a sequential, breadth-first reading... Creating new content, the user establishes a clear, hierarchical directory structure...'\n\nAsks: 'Infer profile for each attribute'",
    "mem0": "System: 'You are analyzing memories extracted by Mem0...'\n\nConverts operations to first-person interactions:\n- I'm reading AAPL_Coverage_Initiation_Notes.md (view #1, 5773 chars)\n- I'm reading AI_Finance_Summit_Attendees.md (view #1, 3195 chars)\n- I created directory 工作概况摘要/附录与参考\n- I wrote file 01_投资分析师工作概况摘要_完整版.md (6479 chars)\n\nAsks: 'Infer profile'",
    "zep": "System: 'You are analyzing a knowledge graph and temporal memory built by Zep...'\n\nGraph facts with temporal ordering:\n- User READ file 'AAPL_Coverage_Initiation_Notes.md' (view #1) [task:T-01]\n- File 'AAPL_Coverage...' REFERENCED 'AI_Finance...' (sequential_access)\n- User CREATED directory '工作概况摘要'\n- User WROTE file '01_投资分析师...md' (6479 chars)\n\nAsks: 'Infer profile'",
    "memos": "System: 'You are analyzing structured memory cells extracted by MemOS...'\n\nHierarchical memory cells with tags:\n- [episodic] Read AAPL_Coverage_Initiation_Notes.md (view #1, 5773 chars)\n- [episodic] Read AI_Finance_Summit_Attendees.md (view #1, 3195 chars)\n- [procedural] Created directory 工作概况摘要/附录与参考\n- [semantic] Wrote file 01_投资分析师...md (6479 chars)\n\nAsks: 'Infer profile'",
    "memu": "System: 'You are analyzing memory units extracted by MemU...'\n\nDiscrete units with type + importance:\n- [access] (imp=0.7) Accessed Tasks.md (view #4)\n- [creation] (imp=0.7) Created/wrote 工作概况摘要/01_...md (6479 chars)\n  Content: # 投资分析师工作概况摘要\n  **分析对象：** Victoria Clarke, CFA...\n\nIncludes actual content of created files.\nAsks: 'Infer profile'",
    "evermemos": "System: 'You are analyzing engram-like memory traces extracted by EverMemOS...'\n\nSegmented by event boundaries:\nSegment T-01_seg1 (boundary: context_switch):\n  [1] FS Snapshot: 26 files...\n  [1] Read file: AAPL_Coverage...\nSegment T-01_seg2 (boundary: context_switch):\n  [1] Cross-file ref: AAPL -> AI_Finance...\n  [1] Read file: AI_Finance...\n\nAsks: 'Infer profile'",
    "filegramos_simple": "System: 'You are an expert at inferring user work habit profiles...'\n\nThree deterministic channels:\n## Channel 1: Procedural Patterns (statistics)\n  reading_strategy: avg_content_length=2580.9, read_ratio=1.0, revisit_ratio=0.558, search_ratio=0.0...\n  output_detail: avg_output_length=3494.7, files_created=3...\n## Channel 2: Semantic Patterns\n  (content-level analysis)\n## Channel 3: Episodic Signatures\n  (cross-session behavioral consistency)\n\nAsks: 'Infer profile'",
}

memory_repr = {
    "full_context": "Raw event sequence (no transformation)",
    "naive_rag": "Chunked text segments + embedding retrieval",
    "eager_summarization": "LLM-generated NL behavioral summary per session",
    "mem0": "First-person interaction memories ('I'm reading X')",
    "zep": "Knowledge graph (entities=files, edges=read/wrote/edited) with temporal metadata",
    "memos": "Hierarchical memory cells tagged as episodic/procedural/semantic",
    "memu": "Typed memory units with importance scores + file content",
    "evermemos": "Engram traces segmented by event boundaries (context switches)",
    "filegramos_simple": "Three-channel deterministic features: procedural stats, semantic patterns, episodic signatures",
}

for i, m in enumerate(METHODS):
    r = row + 1 + i
    style_data_cell(ws4, r, 1, BOLD_FONT).value = METHOD_DISPLAY[m]
    style_data_cell(ws4, r, 2).value = METHOD_DESC[m]
    style_data_cell(ws4, r, 3).value = memory_repr[m]
    style_data_cell(ws4, r, 4).value = prompt_structures[m]
    ws4.row_dimensions[r].height = 120

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 50
ws4.column_dimensions["C"].width = 45
ws4.column_dimensions["D"].width = 70

# ============================================================
# Sheet 5: Attribute Coverage Matrix (tasks × attributes)
# ============================================================
ws5 = wb.create_sheet("Attribute Coverage")

ws5.cell(row=1, column=1, value="Observable Attribute × Task Coverage Matrix").font = Font(
    name="Arial", bold=True, size=12
)

all_attrs = [
    "reading_strategy",
    "output_detail",
    "output_structure",
    "tone",
    "directory_style",
    "naming",
    "edit_strategy",
    "version_strategy",
    "cross_modal_behavior",
]
attr_dims = {
    "reading_strategy": "A (Consumption)",
    "output_detail": "B (Production)",
    "output_structure": "B (Production)",
    "tone": "B (Production)",
    "directory_style": "C (Organization)",
    "naming": "C (Organization)",
    "edit_strategy": "D (Iteration)",
    "version_strategy": "C/D (Org+Iter)",
    "cross_modal_behavior": "F (Cross-Modal)",
}

row = 3
headers5 = ["Attribute", "Dimension"] + [t["id"] for t in tasks] + ["Count"]
for col, h in enumerate(headers5, 1):
    ws5.cell(row=row, column=col, value=h)
style_header_row(ws5, row, len(headers5))

for i, attr in enumerate(all_attrs):
    r = row + 1 + i
    style_data_cell(ws5, r, 1, BOLD_FONT).value = attr
    style_data_cell(ws5, r, 2).value = attr_dims[attr]
    cnt = 0
    for j, t in enumerate(tasks):
        cell = style_data_cell(ws5, r, 3 + j, CHECK_FONT, CENTER_ALIGN)
        if attr in t["attrs"]:
            cell.value = "✓"
            cnt += 1
    style_data_cell(ws5, r, 3 + len(tasks), BOLD_FONT, CENTER_ALIGN).value = cnt

ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 18
for i in range(3, 3 + len(tasks)):
    ws5.column_dimensions[get_column_letter(i)].width = 8
ws5.column_dimensions[get_column_letter(3 + len(tasks))].width = 8

# ============================================================
# Sheet 6: Feature Comparison (p1 vs p2 vs p3 filegramos_simple)
# ============================================================
ws6 = wb.create_sheet("Feature Comparison")

ws6.cell(
    row=1, column=1, value="FileGramOS Channel 1 Features: p1 vs p2 vs p3 (with Ground Truth L/M/R Indicators)"
).font = Font(name="Arial", bold=True, size=12)

# Key features extracted from the three filegramos_simple prompts
feature_sections = [
    (
        "Dim A: Consumption Pattern",
        "reading_strategy",
        [
            ("Ground Truth L/M/R", "p1: A=L (sequential)", "p2: A=L (sequential)", "p3: A=M (targeted)", ""),
            ("read_ratio", "1.000", "1.000", "0.902", "A=L: >=0.7 ✓ | A=M: search present"),
            ("search_ratio", "0.000", "0.000", "0.098", "A=L: =0 ✓ | A=M: >=0.3 (p3 low)"),
            ("total_searches", "0", "0", "4", "p3 uses search, p1/p2 don't"),
            ("browse_ratio", "0.000", "0.000", "0.000", "All zero — no glob/ls browsing"),
            ("revisit_ratio", "0.558", "0.689", "0.730", "All high revisits (unexpected for p3)"),
            ("total_reads", "43", "61", "37", "p2 most reads, p3 fewest"),
            ("unique_files_read", "19", "19", "10", "p3 reads fewer unique files ✓"),
            ("avg_content_length", "2580.9", "2545.6", "2166", "p3 shorter reads"),
            ("context_switch_rate", "0.977", "0.984", "0.973", "Similar across all"),
        ],
    ),
    (
        "Dim B: Production Style",
        "output_detail + output_structure + tone",
        [
            ("Ground Truth L/M/R", "p1: B=L (comprehensive)", "p2: B=L (comprehensive)", "p3: B=R (minimal)", ""),
            ("total_output_chars", "10484", "18431", "2109", "p2>>p1>>p3 ✓✓ (p3 is minimal)"),
            ("avg_output_length", "3494.7", "6143.7", "2109", "p3 much shorter ✓"),
            ("files_created", "3", "3", "1", "p3 single file ✓ (B=R: 1 file)"),
            ("max_output_length", "6479", "10164", "2109", ""),
            ("heading_count", "44", "68", "7", "p3 minimal headings ✓✓"),
            ("heading_max_depth", "4 (####)", "4 (####)", "2 (##)", "p3 shallow ✓ (B=R: at most ##)"),
            ("markdown_table_rows", "128", "283", "0", "p3 no tables ✓✓"),
            ("has_markdown_tables", "True", "True", "False", ""),
            ("list_item_count", "57", "53", "21", ""),
            ("total_words", "1384", "2782", "317", "p3 极简 ✓"),
            ("prose_to_structure_ratio", "0.168", "0.086", "0.143", "p2 most structured"),
        ],
    ),
    (
        "Dim C: Organization Preference",
        "directory_style + naming",
        [
            ("Ground Truth L/M/R", "p1: C=L (deeply_nested)", "p2: C=R (flat)", "p3: C=R (flat)", ""),
            ("dirs_created", "1", "0", "0", "p1 creates dirs, p2/p3 don't ✓"),
            ("max_dir_depth", "2", "0", "0", "p1 nested ✓, p2/p3 flat ✓"),
            ("final_fs_max_depth", "3", "1", "1", "p1 deep hierarchy ✓"),
            ("files_moved", "0", "0", "0", "None moved files"),
            ("avg_filename_length", "12", "5.3", "7", "p1 longest names"),
            ("avg_word_count (filename)", "2", "1", "1", "p1 descriptive naming"),
            ("has_numeric_prefix", "True", "False", "False", "p1 uses 01_, 02_ prefixes"),
            ("has_underscores", "True", "False", "False", "p1 uses underscores"),
        ],
    ),
    (
        "Dim D: Iteration Strategy",
        "edit_strategy + version_strategy",
        [
            ("Ground Truth L/M/R", "p1: D=L (incremental)", "p2: D=R (rewrite)", "p3: D=R (rewrite)", ""),
            ("total_edits", "1", "0", "0", "p1 does incremental edit ✓"),
            ("avg_lines_changed", "2", "0", "0", "p1 small edit ✓"),
            ("small_edit_ratio", "1.000", "0", "0", "p1 all edits are small ✓"),
            ("backup_copies", "0", "0", "0", "None do backups"),
            ("total_overwrites", "0", "0", "0", ""),
            ("total_deletes", "0", "0", "0", ""),
            (
                "NOTE",
                "p1 edits existing file",
                "p2 writes once, no edit",
                "p3 writes once, no edit",
                "D hard to distinguish in Create tasks",
            ),
        ],
    ),
    (
        "Dim F: Cross-Modal Behavior",
        "cross_modal_behavior",
        [
            ("Ground Truth L/M/R", "p1: F=M (balanced)", "p2: F=M (balanced)", "p3: F=R (text-only)", ""),
            ("has_tables_or_data", "True", "True", "False", "p3 pure text ✓"),
            ("markdown_table_rows", "128", "283", "0", "p3 no tables ✓"),
            ("has_images", "False", "False", "False", "None create images"),
            ("image_files_created", "0", "0", "0", ""),
            ("supplementary_ratio", "0.000", "0.000", "0.000", ""),
        ],
    ),
]

row = 3
for section_name, attr_name, features in feature_sections:
    # Section header
    ws6.cell(row=row, column=1, value=section_name).font = Font(name="Arial", bold=True, size=11, color="2F5496")
    ws6.cell(row=row, column=2, value=f"(attributes: {attr_name})").font = Font(
        name="Arial", italic=True, size=9, color="808080"
    )
    row += 1

    # Column headers
    feat_headers = ["Feature", "p1_methodical", "p2_thorough_reviser", "p3_efficient_executor", "L/M/R Indicator Check"]
    for col, h in enumerate(feat_headers, 1):
        ws6.cell(row=row, column=col, value=h)
    style_header_row(ws6, row, len(feat_headers))
    row += 1

    for feat in features:
        for col, val in enumerate(feat):
            cell = style_data_cell(ws6, row, col + 1)
            cell.value = val
            if col == 0:
                cell.font = (
                    BOLD_FONT
                    if not feat[0].startswith("Ground") and not feat[0].startswith("NOTE")
                    else Font(name="Arial", bold=True, size=10, color="C00000")
                )
        row += 1
    row += 1  # gap between sections

ws6.column_dimensions["A"].width = 25
ws6.column_dimensions["B"].width = 25
ws6.column_dimensions["C"].width = 25
ws6.column_dimensions["D"].width = 25
ws6.column_dimensions["E"].width = 50

# ============================================================
# Sheet 7: Per-Dimension Scores
# ============================================================
ws7 = wb.create_sheet("Dimension Scores")

ws7.cell(row=1, column=1, value="Per-Dimension Average Scores by Baseline Method").font = Font(
    name="Arial", bold=True, size=12
)
profile_list_str = "+".join(p.split("_")[0] for p in loaded_profiles)
ws7.cell(
    row=2, column=1, value=f"Attributes mapped to dimensions. Scores averaged across {profile_list_str}."
).font = Font(name="Arial", italic=True, size=9, color="808080")

# Attribute-to-dimension mapping for scoring
ATTR_DIM_MAP = {
    "General": ["name", "role", "language", "error_handling"],
    "A: Consumption": ["reading_strategy"],
    "B: Production": ["tone", "output_detail", "output_structure", "working_style", "thoroughness", "documentation"],
    "C: Organization": ["directory_style", "naming"],
    "D: Iteration": ["edit_strategy", "version_strategy"],
    "F: Cross-Modal": ["cross_modal_behavior"],
}

# Build ALL_SCORES dynamically from results.json
ALL_SCORES = {}
for profile, data in profiles_data.items():
    ALL_SCORES[profile] = {}
    for m in METHODS:
        if m in data and "judge_scores" in data[m]:
            scores = data[m]["judge_scores"]["scores"]
            ALL_SCORES[profile][m] = {attr: scores[attr]["score"] for attr in ATTRIBUTES if attr in scores}
        else:
            ALL_SCORES[profile][m] = {}


# Compute per-dimension average scores
def dim_avg(profile, method, attrs):
    scores = ALL_SCORES.get(profile, {}).get(method, {})
    vals = [scores[a] for a in attrs if a in scores]
    return round(sum(vals) / len(vals), 2) if vals else 0


row = 4
dims_list = list(ATTR_DIM_MAP.keys())

# --- Table 1: averaged across all profiles ---
ws7.cell(row=row, column=1, value=f"Cross-Profile Average ({profile_list_str})").font = Font(
    name="Arial", bold=True, size=11, color="2F5496"
)
row += 1
dim_headers = ["Method"] + dims_list + ["Overall"]
for col, h in enumerate(dim_headers, 1):
    ws7.cell(row=row, column=col, value=h)
style_header_row(ws7, row, len(dim_headers))
row += 1

for m in sorted(METHODS, key=lambda m: method_avgs[m], reverse=True):
    style_data_cell(ws7, row, 1, BOLD_FONT).value = METHOD_DISPLAY[m]
    all_dim_scores = []
    for j, dim_name in enumerate(dims_list):
        attrs = ATTR_DIM_MAP[dim_name]
        avg_val = sum(dim_avg(p, m, attrs) for p in ALL_SCORES) / num_profiles
        avg_val = round(avg_val, 2)
        all_dim_scores.append(avg_val)
        cell = style_data_cell(ws7, row, 2 + j, NORMAL_FONT, CENTER_ALIGN)
        cell.value = avg_val
        cell.fill = score_fill(avg_val)
        cell.number_format = "0.00"
    # Overall
    overall = round(sum(all_dim_scores) / len(all_dim_scores), 2)
    cell = style_data_cell(ws7, row, 2 + len(dims_list), BOLD_FONT, CENTER_ALIGN)
    cell.value = overall
    cell.fill = score_fill(overall)
    cell.number_format = "0.00"
    row += 1

row += 2

# --- Per-profile dimension scores ---
for profile_name in loaded_profiles:
    ws7.cell(row=row, column=1, value=f"Per-Dimension Scores: {profile_name}").font = Font(
        name="Arial", bold=True, size=11, color="2F5496"
    )
    row += 1

    # Attribute mapping note
    ws7.cell(row=row, column=1, value="Attrs in dimension:").font = Font(
        name="Arial", italic=True, size=9, color="808080"
    )
    for j, dim_name in enumerate(dims_list):
        ws7.cell(row=row, column=2 + j, value=", ".join(ATTR_DIM_MAP[dim_name])).font = Font(
            name="Arial", italic=True, size=8, color="808080"
        )
        ws7.cell(row=row, column=2 + j).alignment = WRAP_ALIGN
    row += 1

    for col, h in enumerate(dim_headers, 1):
        ws7.cell(row=row, column=col, value=h)
    style_header_row(ws7, row, len(dim_headers))
    row += 1

    for m in METHODS:
        style_data_cell(ws7, row, 1, BOLD_FONT).value = METHOD_DISPLAY[m]
        dim_scores = []
        for j, dim_name in enumerate(dims_list):
            attrs = ATTR_DIM_MAP[dim_name]
            avg_val = dim_avg(profile_name, m, attrs)
            dim_scores.append(avg_val)
            cell = style_data_cell(ws7, row, 2 + j, NORMAL_FONT, CENTER_ALIGN)
            cell.value = avg_val
            cell.fill = score_fill(avg_val)
            cell.number_format = "0.00"
        overall = round(sum(dim_scores) / len(dim_scores), 2)
        cell = style_data_cell(ws7, row, 2 + len(dims_list), BOLD_FONT, CENTER_ALIGN)
        cell.value = overall
        cell.fill = score_fill(overall)
        cell.number_format = "0.00"
        row += 1
    row += 2

ws7.column_dimensions["A"].width = 22
for i in range(2, 2 + len(dims_list) + 1):
    ws7.column_dimensions[get_column_letter(i)].width = 18

# ============================================================
# Sheet 7b: Channel Analysis
# ============================================================
ws7b = wb.create_sheet("Channel Analysis")

ws7b.cell(row=1, column=1, value="Channel Analysis: Procedural vs Semantic vs Mixed").font = Font(
    name="Arial", bold=True, size=12
)
ws7b.cell(
    row=2, column=1, value=f"Per-attribute scores averaged across all loaded profiles ({profile_list_str})."
).font = Font(name="Arial", italic=True, size=9, color="808080")

PROCEDURAL_ATTRS = [
    "working_style",
    "thoroughness",
    "error_handling",
    "reading_strategy",
    "directory_style",
    "edit_strategy",
    "version_strategy",
    "output_detail",
]
SEMANTIC_ATTRS = ["name", "role", "language", "tone", "output_structure", "documentation"]
MIXED_ATTRS = ["naming", "cross_modal_behavior"]


def channel_avg(method, attr_list):
    """Average score for a list of attributes across all profiles."""
    vals = []
    for p in loaded_profiles:
        p_scores = ALL_SCORES.get(p, {}).get(method, {})
        for a in attr_list:
            if a in p_scores:
                vals.append(p_scores[a])
    return round(sum(vals) / len(vals), 2) if vals else 0


row = 4
ws7b.cell(row=row, column=1, value="Attribute Groupings:").font = Font(name="Arial", bold=True, size=10, color="2F5496")
row += 1
ws7b.cell(row=row, column=1, value="Procedural:").font = BOLD_FONT
ws7b.cell(row=row, column=2, value=", ".join(PROCEDURAL_ATTRS)).font = NORMAL_FONT
row += 1
ws7b.cell(row=row, column=1, value="Semantic:").font = BOLD_FONT
ws7b.cell(row=row, column=2, value=", ".join(SEMANTIC_ATTRS)).font = NORMAL_FONT
row += 1
ws7b.cell(row=row, column=1, value="Mixed:").font = BOLD_FONT
ws7b.cell(row=row, column=2, value=", ".join(MIXED_ATTRS)).font = NORMAL_FONT
row += 2

# Channel comparison table
ch_headers = ["Method", "Procedural Avg", "Semantic Avg", "Mixed Avg", "Overall", "Gap (Proc-Sem)"]
for col, h in enumerate(ch_headers, 1):
    ws7b.cell(row=row, column=col, value=h)
style_header_row(ws7b, row, len(ch_headers))
row += 1

for m in sorted(METHODS, key=lambda m: method_avgs[m], reverse=True):
    style_data_cell(ws7b, row, 1, BOLD_FONT).value = METHOD_DISPLAY[m]
    proc_avg = channel_avg(m, PROCEDURAL_ATTRS)
    sem_avg = channel_avg(m, SEMANTIC_ATTRS)
    mix_avg = channel_avg(m, MIXED_ATTRS)
    overall = round(
        (proc_avg * len(PROCEDURAL_ATTRS) + sem_avg * len(SEMANTIC_ATTRS) + mix_avg * len(MIXED_ATTRS))
        / len(ATTRIBUTES),
        2,
    )
    gap = round(proc_avg - sem_avg, 2)

    for j, val in enumerate([proc_avg, sem_avg, mix_avg, overall, gap]):
        cell = style_data_cell(ws7b, row, 2 + j, NORMAL_FONT, CENTER_ALIGN)
        cell.value = val
        cell.number_format = "0.00"
        if j < 4:
            cell.fill = score_fill(val)
        else:
            # Gap column: green if positive (Procedural > Semantic), red if negative
            if val > 0:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(name="Arial", size=10, color="006100")
            elif val < 0:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(name="Arial", size=10, color="9C0006")
            else:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
    row += 1

row += 2

# Per-profile channel breakdown
ws7b.cell(row=row, column=1, value="Per-Profile Channel Breakdown").font = Font(
    name="Arial", bold=True, size=11, color="2F5496"
)
row += 1

for profile_name in loaded_profiles:
    ws7b.cell(row=row, column=1, value=f"{PROFILE_DISPLAY.get(profile_name, profile_name)}").font = Font(
        name="Arial", bold=True, size=10, color="2F5496"
    )
    row += 1
    for col, h in enumerate(ch_headers, 1):
        ws7b.cell(row=row, column=col, value=h)
    style_header_row(ws7b, row, len(ch_headers))
    row += 1

    for m in METHODS:
        style_data_cell(ws7b, row, 1, BOLD_FONT).value = METHOD_DISPLAY[m]
        p_scores_m = ALL_SCORES.get(profile_name, {}).get(m, {})
        proc_vals = [p_scores_m[a] for a in PROCEDURAL_ATTRS if a in p_scores_m]
        sem_vals = [p_scores_m[a] for a in SEMANTIC_ATTRS if a in p_scores_m]
        mix_vals = [p_scores_m[a] for a in MIXED_ATTRS if a in p_scores_m]
        proc_avg = round(sum(proc_vals) / len(proc_vals), 2) if proc_vals else 0
        sem_avg = round(sum(sem_vals) / len(sem_vals), 2) if sem_vals else 0
        mix_avg = round(sum(mix_vals) / len(mix_vals), 2) if mix_vals else 0
        overall = (
            round(
                (sum(proc_vals) + sum(sem_vals) + sum(mix_vals)) / (len(proc_vals) + len(sem_vals) + len(mix_vals)), 2
            )
            if (proc_vals or sem_vals or mix_vals)
            else 0
        )
        gap = round(proc_avg - sem_avg, 2)

        for j, val in enumerate([proc_avg, sem_avg, mix_avg, overall, gap]):
            cell = style_data_cell(ws7b, row, 2 + j, NORMAL_FONT, CENTER_ALIGN)
            cell.value = val
            cell.number_format = "0.00"
            if j < 4:
                cell.fill = score_fill(val)
            else:
                if val > 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(name="Arial", size=10, color="006100")
                elif val < 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(name="Arial", size=10, color="9C0006")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
        row += 1
    row += 2

ws7b.column_dimensions["A"].width = 22
ws7b.column_dimensions["B"].width = 16
ws7b.column_dimensions["C"].width = 16
ws7b.column_dimensions["D"].width = 14
ws7b.column_dimensions["E"].width = 14
ws7b.column_dimensions["F"].width = 16

# ============================================================
# Per-Profile Full View sheets (Prompt + Inference + Score + Justification)
# One sheet per profile that has data. Each method = one vertical block:
#   1) Method header + avg score + prompt length
#   2) Full prompt text
#   3) Attribute table: Attribute | Ground Truth | Inferred | Score | Justification
# ============================================================

INF_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green for inferred
JUST_FILL = PatternFill("solid", fgColor="F2F2F2")  # light gray for justification
GT_FILL = PatternFill("solid", fgColor="D6E4F0")  # light blue for ground truth
PROMPT_FILL = PatternFill("solid", fgColor="FFF8E7")  # light yellow for prompt
METHOD_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
METHOD_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=13)

for profile_key in loaded_profiles:
    if profile_key not in profiles_data:
        continue

    data = profiles_data[profile_key]
    short = profile_key.split("_")[0].upper()  # P1, P2, P3
    ws = wb.create_sheet(f"{short} Full")

    # Title
    ws.cell(
        row=1, column=1, value=f"{PROFILE_DISPLAY[profile_key]} — Input Prompt + Inference Results + Ground Truth"
    ).font = Font(name="Arial", bold=True, size=14)
    ws.cell(
        row=2,
        column=1,
        value="Each method block: input prompt → per-attribute comparison (GT vs Inferred vs Score vs Justification)",
    ).font = Font(name="Arial", italic=True, size=9, color="808080")

    # Get ground truth (same across methods)
    gt = None
    for m in METHODS:
        if m in data and "ground_truth" in data[m]:
            gt = data[m]["ground_truth"]
            break

    row = 4

    for m_idx, m in enumerate(METHODS):
        if m not in data:
            continue

        method_data = data[m]
        avg_score = method_data.get("avg_score", 0)
        prompt_length = method_data.get("prompt_length", 0)

        # ---- Method header bar (full width, colored) ----
        header_text = f"▌ {m_idx + 1}/{len(METHODS)}  {METHOD_DISPLAY[m]}    |    Avg Score: {round(avg_score, 2)}    |    Prompt: {prompt_length} chars"
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = METHOD_HEADER_FILL
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1, value=header_text).font = METHOD_HEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 30
        row += 1

        # ---- Prompt section ----
        ws.cell(row=row, column=1, value="INPUT PROMPT (Memory Representation):").font = Font(
            name="Arial", bold=True, size=10, color="996600"
        )
        row += 1

        # Read prompt file
        prompt_path = BENCH_DIR / profile_key / f"{m}_prompt.txt"
        prompt_text = ""
        if prompt_path.exists():
            with open(prompt_path, encoding="utf-8") as f:
                prompt_text = f.read()

        # Put prompt in a merged cell spanning columns A-E
        prompt_cell = ws.cell(row=row, column=1)
        prompt_cell.value = prompt_text[:32000] if prompt_text else "(no prompt file)"
        prompt_cell.font = Font(name="Consolas", size=9)
        prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")
        prompt_cell.fill = PROMPT_FILL
        prompt_cell.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        # Dynamic row height based on line count (capped)
        line_count = min(prompt_text.count("\n") + 1, 200)
        ws.row_dimensions[row].height = max(80, min(line_count * 13, 600))
        row += 1

        # ---- Results table header ----
        row += 1
        result_headers = ["Attribute", "Ground Truth", "Inferred", "Score", "Justification"]
        for col, h in enumerate(result_headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header_row(ws, row, len(result_headers))
        row += 1

        # ---- Per-attribute rows ----
        for attr in ATTRIBUTES:
            # Attribute name
            style_data_cell(ws, row, 1, BOLD_FONT).value = attr

            # Ground truth
            gt_cell = style_data_cell(ws, row, 2)
            gt_cell.value = gt.get(attr, "-") if gt else "-"
            gt_cell.fill = GT_FILL

            # Inferred
            inferred = method_data.get("inferred_profile", {}).get(attr, "-")
            inf_cell = style_data_cell(ws, row, 3)
            inf_cell.value = inferred
            inf_cell.fill = INF_FILL

            # Score + Justification
            scores_data = method_data.get("judge_scores", {}).get("scores", {})
            if attr in scores_data:
                s = scores_data[attr]["score"]
                just = scores_data[attr].get("justification", "")
            else:
                s = "-"
                just = ""

            score_cell = style_data_cell(ws, row, 4, BOLD_FONT, CENTER_ALIGN)
            score_cell.value = s
            if isinstance(s, (int, float)):
                score_cell.fill = score_fill(s)

            just_cell = style_data_cell(ws, row, 5)
            just_cell.value = just
            just_cell.fill = JUST_FILL

            row += 1

        # ---- Overall avg row ----
        style_data_cell(ws, row, 1, BOLD_FONT).value = "OVERALL"
        style_data_cell(ws, row, 2).value = "-"
        style_data_cell(ws, row, 3).value = "-"
        avg_cell = style_data_cell(ws, row, 4, BOLD_FONT, CENTER_ALIGN)
        avg_cell.value = round(avg_score, 2)
        avg_cell.fill = score_fill(avg_score)
        avg_cell.number_format = "0.00"
        style_data_cell(ws, row, 5).value = f"Average across {len(ATTRIBUTES)} attributes"
        row += 1

        # ---- Gap between methods ----
        row += 2

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 65

# ============================================================
# Save
# ============================================================
output_path = "/Users/choiszt/Desktop/code/Synvo/FileGram/bench/FileGram_Bench_Report.xlsx"
wb.save(output_path)
print(f"Excel saved to: {output_path}")
